from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook


def _cell(row, index, default=""):
    if index >= len(row):
        return default
    value = row[index]
    if value is None:
        return default
    return str(value).strip()


def _headers_map(header_row):
    mapping = {}
    for idx, raw in enumerate(header_row):
        if raw is None:
            continue
        key = str(raw).strip().replace("\xa0", " ")
        mapping[key] = idx
    return mapping


def _get(mapping, row, *names, default=""):
    for name in names:
        if name in mapping:
            return _cell(row, mapping[name], default=default)
    return default


def iter_xlsx_rows(file_bytes: bytes):
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return
        mapping = _headers_map(header)
        for row_number, row in enumerate(rows, start=2):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            yield row_number, mapping, row
    finally:
        workbook.close()


def parse_decimal(value, default="0"):
    text = str(value or default).strip().replace(",", "")
    if not text:
        text = default
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def parse_date(value):
    if value is None or value == "":
        return None
    if hasattr(value, "date") and not isinstance(value, str):
        return value.date() if hasattr(value, "hour") else value
    text = str(value).strip()
    if not text:
        return None
    from datetime import datetime

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            continue
    return None


def parse_datetime(value):
    if value is None or value == "":
        return None
    if hasattr(value, "hour") and not isinstance(value, str):
        return value
    text = str(value).strip()
    if not text:
        return None
    from datetime import datetime

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    return None


def map_supplier_row(mapping, row):
    return {
        "gln": _get(mapping, row, "کد GLN", "GLN"),
        "name": _get(mapping, row, "نام تامین کننده", "نام تأمین کننده"),
        "country": _get(mapping, row, "کشور"),
        "province": _get(mapping, row, "استان"),
        "county": _get(mapping, row, "شهرستان"),
        "city": _get(mapping, row, "شهر"),
        "address": _get(mapping, row, "آدرس"),
        "ceo_name": _get(mapping, row, "نام مدیر  عامل", "نام مدیر عامل"),
        "company_type": _get(mapping, row, "نوع شرکت"),
        "national_id": _get(mapping, row, "شناسه شرکت", "کد ملی"),
        "postal_code": _get(mapping, row, "کد پستی"),
        "phone": _get(mapping, row, "تلفن"),
        "ceo_national_id": _get(mapping, row, "مدیر عامل کد ملی"),
        "technical_name": _get(mapping, row, "نام مسئول فنی"),
        "technical_national_id": _get(
            mapping, row, "شناسه ملی مسئول فنی", ".شناسه ملی مسئول فنی"
        ),
        "mobile": _get(mapping, row, "موبایل"),
        "email": _get(mapping, row, "ایمیل"),
    }


def map_pharmacy_row(mapping, row):
    name = _get(mapping, row, "نام داروخانه")
    customer_national_id = _get(mapping, row, "شناسه ملی مشتری")
    pharmacy_type = _get(mapping, row, "نوع داروخانه")
    if name.isdigit() and pharmacy_type and not pharmacy_type.isdigit():
        customer_national_id = customer_national_id or name
        name = pharmacy_type
        pharmacy_type = ""
    return {
        "gln": _get(mapping, row, "کد GLN", "GLN"),
        "university_name": _get(mapping, row, "نام دانشگاه"),
        "service_type": _get(mapping, row, "نوع سرویس داروخانه"),
        "pharmacy_type": pharmacy_type,
        "name": name,
        "customer_national_id": customer_national_id,
        "postal_code": _get(mapping, row, "کد پستی"),
        "owner_name": _get(mapping, row, "نام مالک / رابط مرکز"),
        "responsible_national_id": _get(mapping, row, "کد ملی مسئول"),
        "founder_mobile": _get(mapping, row, "شماره همراه موسس"),
        "landline": _get(mapping, row, "شماره ثابت"),
        "address": _get(mapping, row, "آدرس"),
        "province": _get(mapping, row, "استان"),
        "county": _get(mapping, row, "شهرستان"),
        "city": _get(mapping, row, "شهر"),
        "national_id": customer_national_id
        or _get(mapping, row, "کد ملی مسئول")
        or _get(mapping, row, "کد GLN", "GLN"),
    }


def map_distributor_row(mapping, row):
    return {
        "name": _get(mapping, row, "نام شرکت"),
        "gln": _get(mapping, row, "GLN", "کد GLN"),
        "national_id": _get(mapping, row, "کد ملی", "شناسه شرکت"),
        "ceo_name": _get(mapping, row, "نام مدیر  عامل", "نام مدیر عامل"),
        "company_type": _get(mapping, row, "اطلاعات  شرکت ها.نوع شرکت", "نوع شرکت"),
        "country": _get(mapping, row, "اطلاعات  شرکت ها.کشور", "کشور"),
        "province": _get(mapping, row, "اطلاعات  شرکت ها.استان", "استان"),
        "postal_code": _get(mapping, row, "کد پستی"),
        "address": _get(mapping, row, "اطلاعات  شرکت ها.آدرس", "آدرس"),
        "phone": _get(mapping, row, "اطلاعات  شرکت ها.تلفن", "تلفن"),
        "ceo_national_id": _get(mapping, row, "مدیر عامل کد ملی"),
        "technical_name": _get(mapping, row, "نام مسئول فنی"),
        "technical_national_id": _get(
            mapping, row, ".شناسه ملی مسئول فنی", "شناسه ملی مسئول فنی"
        ),
        "mobile": _get(mapping, row, "موبایل"),
        "email": _get(mapping, row, "ایمیل"),
        "branch_gln": _get(mapping, row, "GLN شعبه توزیع کننده"),
        "branch_name": _get(mapping, row, "نام شعبه توزیع کننده"),
        "branch_postal_code": _get(mapping, row, "کد پستی شعبه"),
        "branch_province": _get(mapping, row, "لیست توزیع کنندگان.استان"),
        "branch_county": _get(mapping, row, "لیست توزیع کنندگان.شهرستان"),
        "branch_city": _get(mapping, row, "لیست توزیع کنندگان.شهر"),
        "branch_address": _get(mapping, row, "لیست توزیع کنندگان.آدرس"),
    }


def map_medicine_row(mapping, row):
    return {
        "external_id": _get(mapping, row, "شناسه"),
        "name": _get(mapping, row, "نام"),
        "strength": _get(mapping, row, "قدرت دارویی(فرمت قدیمی)", "قدرت دارویی"),
        "molecule": _get(mapping, row, "مولکول"),
        "route": _get(mapping, row, "نحوه مصرف"),
        "dosage_form": _get(mapping, row, "شکل دارویی"),
        "atc_code": _get(mapping, row, "کد ATC"),
        "formulary_code": _get(mapping, row, "کد فهرست"),
        "access_level": _get(mapping, row, "سطح دسترسی"),
        "drug_type": _get(mapping, row, "نوع دارو"),
        "clinical_use": _get(mapping, row, "کاربرد بالینی تایید شده"),
        "formulary_entry_date": _get(mapping, row, "تاریخ کارگروه ورود به فهرست"),
    }


def iter_ods_rows(file_bytes: bytes):
    """Parse first sheet of an ODS workbook into header-mapped rows."""
    from odf.opendocument import load
    from odf.table import Table, TableCell, TableRow
    from odf.text import P

    document = load(BytesIO(file_bytes))
    tables = document.getElementsByType(Table)
    if not tables:
        return

    def cell_text(cell):
        return "".join(str(p) for p in cell.getElementsByType(P)).strip()

    raw_rows = []
    for table_row in tables[0].getElementsByType(TableRow):
        values = []
        for cell in table_row.getElementsByType(TableCell):
            attrs = getattr(cell, "attributes", {}) or {}
            repeat = 1
            for key, val in attrs.items():
                if "numbercolumnsrepeated" in str(key).lower():
                    repeat = int(val or 1)
                    break
            text = cell_text(cell)
            values.extend([text] * min(repeat, 40))
        raw_rows.append(values)

    if not raw_rows:
        return
    mapping = _headers_map(raw_rows[0])
    for row_number, row in enumerate(raw_rows[1:], start=2):
        if not any(str(c).strip() for c in row):
            continue
        yield row_number, mapping, row


def map_insurance_price_row(mapping, row):
    return {
        "generic_code": _get(mapping, row, "generic"),
        "generic_name": _get(mapping, row, "generic_name"),
        "insurance_price": _get(mapping, row, "insurance_price"),
        "subsidy_price": _get(mapping, row, "subsidy_price"),
        "insurance_type": _get(mapping, row, "insurance_type"),
        "letter_shamsi_date": _get(mapping, row, "letter_shamsi_date"),
        "letter_miladi_date": _get(mapping, row, "letter_miladi_date"),
        "package_number": _get(mapping, row, "package_number"),
        "last_update_date": _get(mapping, row, "last_update_date"),
        "source_created_at": _get(mapping, row, "created_at"),
        "source_updated_at": _get(mapping, row, "updated_at"),
    }
